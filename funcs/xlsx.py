from openpyxl import load_workbook

# Caminho da planilha
caminho_planilha = "funcs/planilha.xlsx"


# Carregar planilha
def carregar_planilha():
    """Carrega a planilha e retorna a planilha e a aba ativa."""
    try:
        wb = load_workbook(caminho_planilha)
        ws = wb.active
        return wb, ws
    except Exception as e:
        print(f"Erro ao carregar a planilha: {str(e)}")
        return None, None

def verificar_oculto(ws):
    """Verifica na coluna H da planilha de controle se a página deve ser oculta."""
    try:
        valor_h4 = ws["H8"].value
        if valor_h4 == "Oculta":
            print("A página deve ser oculta.")
            return True
        else:
            print("A página não deve ser oculta.")
            return False
    except Exception as e:
        print(f"Erro ao verificar se a página deve ser oculta: {str(e)}")
        return False

def pegar_conteudo_input_por_tag_e_class(ws):
    """Obtém o conteúdo de um campo de entrada (input) usando tags e classes da planilha."""
    try:
        # Localiza o valor de uma célula específica na planilha, por exemplo, P4
        valor = ws["P4"].value
        print(f"Valor da célula P4: {valor}")
        return valor
    except Exception as e:
        print(f"Erro ao pegar o conteúdo da célula: {str(e)}")
        return None

def pegar_conteudo_input_por_id(ws):
    """Obtém o conteúdo de um campo de entrada (input) pelo ID da planilha."""
    try:
        # Localiza o valor de uma célula específica na planilha, por exemplo, P4
        valor = ws["P4"].value
        print(f"Valor da célula P4: {valor}")
        return valor
    except Exception as e:
        print(f"Erro ao pegar o conteúdo da célula: {str(e)}")
        return None

def verificar_e_filtrar_tipo_pagina(ws):
    """Verifica o tipo de página nas células da coluna P, filtra o texto após ': ', remove páginas vinculadas e valores '-', e retorna um dicionário com os itens e seus respectivos status de visibilidade no formato 'visibilidade': ['tipo de pagina']."""
    try:
        result = {}  # Dicionário para armazenar os resultados
        max_linhas = ws.max_row  # Número máximo de linhas da planilha
        
        for linha in range(4, max_linhas + 1):  # Itera da linha 4 até o final
            tipo_pagina = ws[f"P{linha}"].value  # Pega o valor da célula na coluna P da linha atual (tipo de página)
            visibilidade = ws[f"H{linha}"].value  # Pega o valor da célula na coluna H da linha atual (status de visibilidade)
            
            if tipo_pagina:  # Verifica se o valor de tipo de página não é None ou vazio
                # Extrai o valor após ": " para o tipo de página
                tipo_extraido = tipo_pagina.split(": ", 1)[-1] if ": " in tipo_pagina else tipo_pagina
                
                # Verifica se não é um valor de página vinculada ou "-"
                if "Vincular a uma página deste site" not in tipo_extraido and tipo_extraido != '-':
                    # Adiciona o tipo de página com seu status de visibilidade
                    if visibilidade == "Oculta":
                        if "Oculto" not in result:
                            result["Oculto"] = []  # Inicializa a lista para "Oculto"
                        result["Oculto"].append(tipo_extraido)
                    elif visibilidade == "Menu":
                        if "Menu" not in result:
                            result["Menu"] = []  # Inicializa a lista para "Menu"
                        result["Menu"].append(tipo_extraido)
                    else:
                        if "undefined" not in result:
                            result["undefined"] = []  # Inicializa a lista para "undefined"
                        result["undefined"].append(tipo_extraido)  # Para outros valores ou células vazias
                else:
                    if "undefined" not in result:
                        result["undefined"] = []  # Inicializa a lista para "undefined"
                    result["undefined"].append("undefined-page")  # Marca como "undefined-page" se for "-"
        
        # Retorna o dicionário com os tipos e status das páginas
        return result
    except Exception as e:
        print(f"Erro ao verificar e filtrar os tipos de página: {str(e)}")
        return {}

def pegar_itens_colunas_ph(ws):
    """
    Acessa os itens das colunas P e H a partir da linha 4 até o final da planilha,
    aplicando substituições conforme necessário.

    Regras:
    - Se coluna P == "-", substituir por "Widget".
    - Se coluna H == "-", substituir por "Oculta".

    Parâmetros:
    - ws: Objeto da planilha.

    Retorna:
    - Lista de tuplas (H, P), onde cada tupla contém os valores ajustados das colunas.
    """
    try:
        itens = []  # Lista para armazenar os itens encontrados
        max_linhas = ws.max_row  # Número máximo de linhas da planilha
        
        for linha in range(4, max_linhas + 1):
            valor_p = ws[f"P{linha}"].value  # Obtém o valor da célula na coluna P
            valor_h = ws[f"H{linha}"].value  # Obtém o valor da célula na coluna H
            
            # Substituições conforme a regra definida
            if valor_p == "-":
                valor_p = "Widget"
            if valor_h == "-":
                valor_h = "Oculta"
            
            if valor_p and valor_h:  # Garante que ambos os valores não sejam None
                itens.append((valor_h, valor_p))  # Adiciona os valores à lista
        
        return itens  # Retorna a lista de itens encontrados
    
    except Exception as e:
        print(f"Erro ao acessar os itens das colunas P e H: {str(e)}")
        return []

