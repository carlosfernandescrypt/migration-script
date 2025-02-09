#acessar pasta raiz:

import sys
import os

# Adiciona a pasta raiz ao caminho de pesquisa de módulos
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from config  import BASE_URL, PATH_PLANILHA, PASSWORD, EMAIL


# Configurações
url_base = BASE_URL
caminho_planilha = PATH_PLANILHA
password = PASSWORD
email = EMAIL
# Inicialização do WebDriver
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")
driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, 20)

# Carregar planilha
wb = load_workbook(caminho_planilha)
ws = wb.active

def fazer_login():
    """Realiza o login no sistema."""
    try:
        print("Acessando a página de login...")
        driver.get(url_base)
        
        # Clicar no botão de login
        print("Clicando no botão de login...")
        botao_login = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, ".sign-in.text-default.btn.btn-sm.btn-unstyled")
        ))
        botao_login.click()
        
        # Preencher e-mail
        print("Preenchendo e-mail...")
        campo_email = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, ".field.clearable.form-control")
        ))
        campo_email.clear()
        campo_email.send_keys(email)
        
        # Preencher senha
        print("Preenchendo senha...")
        campo_senha = wait.until(EC.element_to_be_clickable(
            (By.ID, "_com_liferay_login_web_portlet_LoginPortlet_password")
        ))
        campo_senha.clear()
        campo_senha.send_keys(password)
        
        # Clicar no botão de entrar
        print("Clicando no botão de entrar...")
        botao_entrar = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, ".btn.btn-primary")
        ))
        botao_entrar.click()


        # Navegar para a área de administração usando a nova classe
        print("Navegando para área de administração...")
        wait.until(EC.element_to_be_clickable(
            (By.ID, "panel-manage-site_administration_build-link")
        )).click()
        
        # Clicar no link de páginas
        print("Acessando seção de páginas...")
        wait.until(EC.element_to_be_clickable(
            (By.ID, "_com_liferay_product_navigation_product_menu_web_portlet_ProductMenuPortlet_portlet_com_liferay_layout_admin_web_portlet_GroupPagesPortlet")
        )).click()
        
        return True
    except Exception as e:
        print(f"Erro ao fazer login: {str(e)}")
        return False


# Verificar se o localhost está acessível
if not fazer_login():
    print(f"Erro: Não foi possível fazer login em {url_base}. Verifique as credenciais e o servidor.")
    driver.quit()
    exit()




def clicar_botao(css_selector):
    """Clica em um botão específico com base no seletor CSS."""
    try:
        botao = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, css_selector)))
        botao.click()
        time.sleep(1)  # Pequeno delay para garantir a transição
    except Exception as e:
        print(f"Erro ao clicar no botão {css_selector}: {str(e)}")

def clicar_botao_novo():
    """Clica no botão 'Novo' usando seu texto específico."""
    try:
        botao_novo = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Novo' and @title='Novo']//span[text()='Novo']")))
        botao_novo.click()
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao clicar no botão 'Novo': {str(e)}")

def clicar_link_pagina():
    """Clica no link 'Página' dentro do menu suspenso."""
    try:
        link_pagina = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[text()='Página']")))
        link_pagina.click()
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao clicar no link 'Página': {str(e)}")

def clicar_div_pagina_definida():
    """Clica na div 'Página definida'."""
    try:
        div_pagina_definida = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='card-title' and @title='Página definida']//span[text()='Página definida']")))
        div_pagina_definida.click()
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao clicar na div 'Página definida': {str(e)}")

def clicar_div_pagina_widget():
    """Clica na div 'Página Widget'."""
    try:
        div_pagina_widget = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='card-title' and @title='Página de Widget']//span[text()='Página de Widget']")))
        div_pagina_widget.click()
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao clicar na div 'Página definida': {str(e)}")

def clicar_div_vincular_pagina_deste_site():
    """Clica na div 'Vincular Pagina Deste Site'."""
    try:
        clicar_div_vincular_pagina_deste_site = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='card-title' and @title='Vincular a uma página deste site']//span[text()='Vincular a uma página deste site']")))
        clicar_div_vincular_pagina_deste_site.click()
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao clicar na div 'Página definid': {str(e)}")

def clicar_div_vincular_url():
    """Clica na div 'Vincular Url'."""
    try:
        div_vincular_url = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='card-title' and @title='Vincular a uma URL']//span[text()='Vincular a uma URL']")))
        div_vincular_url.click()
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao clicar na div 'Página definida': {str(e)}")

def preencher_input_nome(nome_pagina):
    """Muda para o iframe, espera 1 segundo, clica no campo de nome, preenche com 'Teste' e pressiona Enter."""
    try:
        time.sleep(1)  # Espera antes de executar a função
        iframe = wait.until(EC.presence_of_element_located((By.ID, "addLayoutDialog_iframe_")))
        print("[LOG] Alternando para o iframe...")
        driver.switch_to.frame(iframe)  # Mudar para dentro do iframe
        campo_nome = wait.until(EC.element_to_be_clickable(
            (By.ID, "_com_liferay_layout_admin_web_portlet_GroupPagesPortlet_name")
        ))
        campo_nome.click()  # Garante que o campo seja ativado
        campo_nome.clear()  # Remove qualquer texto anterior
        campo_nome.send_keys(nome_pagina)  # Digita o texto
        campo_nome.send_keys(Keys.RETURN)  # Pressiona Enter para confirmar
        print("[LOG] Campo nome preenchido e Enter pressionado.")
        
        # Remova a espera pela próxima página:
        # print("[LOG] Esperando a próxima página carregar...")
        # wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".some-element-on-next-page")))  # Não é necessário aguardar agora

        # Agora você pode continuar com o restante do processo sem esperar
        # Interação com campos na nova página, se necessário

        # Retornar ao contexto principal
        driver.switch_to.default_content()

    except Exception as e:
        print(f"[ERRO] Falha ao preencher o campo de nome dentro do iframe: {str(e)}")
        driver.switch_to.default_content()  # Retorna ao contexto principal mesmo em caso de erro



def verificar_oculto():
    """Verifica na coluna H da planilha de controle se a página deve ser oculta."""
    try:
        valor_h4 = ws["H4"].value
        if valor_h4 == "Oculta":
            print("A página deve ser oculta.")
            return True
        else:
            print("A página não deve ser oculta.")
            return False
    except Exception as e:
        print(f"Erro ao verificar se a página deve ser oculta: {str(e)}")
        return False

def apertar_enter(campo_input):
    """Simula o pressionamento da tecla Enter no campo de entrada."""
    try:
        campo_input.send_keys(Keys.RETURN)
        print("Tecla Enter pressionada.")
    except Exception as e:
        print(f"Erro ao pressionar Enter: {str(e)}")

def selecionar_pagina_widget():
    """Aguarda e clica no card 'Página de Widget'."""
    try:
        pagina_widget = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//li[@data-qa-id='cardPageItemDirectory']//p[@title='Página de Widget']"
        )))
        pagina_widget.click()
    except Exception as e:
        print(f"Erro ao selecionar 'Página de Widget': {str(e)}")




def selecionar_layout_1_coluna():
    """Muda para o iframe, aguarda e clica no card '1 Coluna'."""
    try:
        card_1_coluna = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//div[contains(@class, 'card-type-template')]//span[@title='1 Coluna']"
        )))
        card_1_coluna.click()

        driver.switch_to.default_content()

    except Exception as e:
        driver.switch_to.default_content()
        print(f"Erro ao selecionar o layout '1 Coluna': {str(e)}")

def pegar_conteudo_input_por_tag_e_class():
    """Obtém o conteúdo de um campo de entrada (input) usando tags e classes."""
    try:
        # Localiza o campo de input usando a classe e a tag
        campo_input = wait.until(EC.presence_of_element_located((
            By.XPATH, "//div[@class='input-group-item']//input[@class='form-control language-value']"
        )))
        
        # Pega o valor do campo de entrada
        valor = campo_input.get_attribute('value')
        print(f"Valor do campo de entrada: {valor}")
        return valor
    except Exception as e:
        print(f"Erro ao pegar o conteúdo do campo de entrada: {str(e)}")
        return None

def pegar_conteudo_input_por_id():
    """Obtém o conteúdo de um campo de entrada (input) pelo ID."""
    try:
        # Localiza o campo de input pelo ID
        campo_input = wait.until(EC.presence_of_element_located((
            By.ID, "_com_liferay_layout_admin_web_portlet_GroupPagesPortlet_friendlyURL"
        )))
        
        # Pega o valor do campo de entrada
        valor = campo_input.get_attribute('value')
        print(f"Valor do campo de entrada: {valor}")
        return valor
    except Exception as e:
        print(f"Erro ao pegar o conteúdo do campo de entrada: {str(e)}")
        return None

def pegar_botao_salvar():
    """Obtém o botão 'Salvar' dentro da div com a classe 'sheet-footer'."""
    try:
        # Localiza o botão "Salvar" usando XPath
        botao_salvar = wait.until(EC.presence_of_element_located((
            By.XPATH, "//div[@class='sheet-footer']//button[@class='btn btn-primary']//span[text()='Salvar']"
        )))
        
        # Clique no botão
        botao_salvar.click()
        print("Botão 'Salvar' clicado com sucesso!")
    except Exception as e:
        print(f"Erro ao localizar ou clicar no botão 'Salvar': {str(e)}")

def clicar_label_por_for(valor_for):
    """Clica em um label baseado no atributo 'for'."""
    try:
        label = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"label[for='{valor_for}']")))
        label.click()
        print(f"Label com 'for'='{valor_for}' clicado com sucesso!")
    except Exception as e:
        print(f"Erro ao clicar no label com 'for'='{valor_for}': {str(e)}")

def selecionar_botao_selecionar():
    """Aguarda e clica no botão 'Selecionar'."""
    try:
        botao_selecionar = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//button[@id='_com_liferay_layout_admin_web_portlet_GroupPagesPortlet_selectLayoutButton'][text()='Selecionar']"
        )))
        botao_selecionar.click()
        print("Botão 'Selecionar' clicado com sucesso.")
    except Exception as e:
        print(f"Erro ao selecionar o botão 'Selecionar': {str(e)}")

def mudar_url(nova_url):
    """Navega para uma nova URL no navegador controlado pelo Selenium."""
    try:
        print(f"Acessando a URL: {nova_url}")
        driver.get(nova_url)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))  # Aguarda o carregamento da página
        print("Página carregada com sucesso!")
    except Exception as e:
        print(f"Erro ao acessar a URL {nova_url}: {str(e)}")



def criar_pagina(type):
    if type == "Definida":
        clicar_div_pagina_definida()
        preencher_input_nome(valor_extraido_pd)
        pegar_conteudo_input_por_id()
        mudar_url("{url_base}/home")
    elif type == "Widget":
        clicar_div_pagina_widget()
        preencher_input_nome(valor_extraido_pw)
        selecionar_layout_1_coluna()
        pegar_conteudo_input_por_id()
        clicar_label_por_for("_com_liferay_layout_admin_web_portlet_GroupPagesPortlet_hidden") # isso aqui server para clicar no botão de ocultar
    elif type == "Vincular a uma página deste site":
        clicar_div_vincular_pagina_deste_site()
        preencher_input_nome(valor_extraido_vpds)
        pegar_conteudo_input_por_id()
        selecionar_botao_selecionar()
    elif type == "Vincular a uma URL":
        clicar_div_vincular_url()
        preencher_input_nome()
    pass

try:

    # clicar no botão class="dropdown-toggle nav-btn btn btn-primary"
    clicar_botao_novo()
    # class="dropdown-item" Página
    clicar_link_pagina()
    #valor_p4 = ws["P4"].value
    #valor_extraido = valor_p4.split(": ", 1)[-1]
    valor_extraido_pd = "Definida"
    valor_extraido_pw = "Widget"
    valor_extraido_vpds = "Vincular a uma página deste site"
    valor_extraido_vurl = "Vincular a uma URL"

    print(f"Valor da célula P4: {valor_extraido_pw}")

    criar_pagina(valor_extraido_pd)

finally:
    #wb.close()
    print("Planilha fechada.")